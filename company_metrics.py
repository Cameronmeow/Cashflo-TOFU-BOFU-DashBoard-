# company_metrics.py
"""
Enrich vendor data with:
  • Cash-Rich Status
  • Indicative Interest Rate (%)
  • Dependency %  + Dependency Slab
  • Deviation vs. industry working-capital benchmarks
  • Industry / Nature-of-Business benchmark table

Returns two DataFrames (enriched, benchmark) for Streamlit
and still offers a CLI that saves a coloured Excel workbook
(“Calculations” + “Industry Benchmarks”).

Usage in Streamlit
------------------
    import company_metrics as cm
    enriched_df, bench_df = cm.enrich_dataframe(df)

CLI
---
    python company_metrics.py          # prompts for folder of XLSX/CSV files
"""

from __future__ import annotations

import glob
import os
import tkinter as tk
from tkinter import filedialog
from typing import List, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ────────────────────────── helpers ──────────────────────────
def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def _get_col(df: pd.DataFrame, name: str) -> str:
    mask = df.columns.str.lower() == name.lower()
    if not mask.any():
        raise KeyError(f"Column '{name}' not found")
    return df.columns[mask][0]


def _safe_series(
    df: pd.DataFrame, col: str, default=np.nan, numeric: bool = True
) -> pd.Series:
    try:
        s = df[_get_col(df, col)]
        return _to_num(s) if numeric else s
    except KeyError:
        return pd.Series(default, index=df.index)


def _parse_slab(txt: str | float) -> float:
    if pd.isna(txt):
        return np.nan
    if isinstance(txt, (int, float)):
        return float(txt)

    txt = (
        str(txt)
        .replace("Rs", "")
        .replace("Cr", "")
        .replace(",", "")
        .strip()
        .lower()
    )
    if "to" in txt:
        lo, hi = [float(x) for x in txt.split("to")]
        return (lo + hi) / 2
    if "and above" in txt:
        return float(txt.split()[0])
    return np.nan


def _fy_label(ts: pd.Timestamp) -> str:
    yr = ts.year + 1 if ts.month >= 4 else ts.year
    return f"FY{str(yr)[-2:]}"


# ────────────────────────── main enricher ──────────────────────────
def enrich_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()  # keep caller’s df intact

    # ---------- Cash-Rich Status ----------
    cash_eq = _safe_series(df, "Cash and Cash Equivalents", 0)
    invest = _safe_series(df, "Current investments", 0)
    st_borr = _safe_series(df, "Short term borrowings").replace(0, np.nan)
    rev_grow = _safe_series(df, "Revenue growth in %", 0)

    rating_col = next(
        (c for c in df.columns if c.lower() in ("latest credit ratings", "rating")), None
    )
    rating = df[rating_col].astype(str) if rating_col else ""

    ratio = (cash_eq + invest) / st_borr
    cash_rich_mask = ((ratio > 2) & (rev_grow < 15)) | rating.str.upper().str.startswith(
        "AA"
    )
    df["Cash-Rich Status"] = np.where(cash_rich_mask, "Cash-Rich", "Non-Cash Rich")

    # ---------- Indicative Interest Rate ----------
    fin_cost_pct = _safe_series(df, "Finance Cost (% of Sales)") / 100
    turnover = _safe_series(df, "Annual Revenue")
    total_debt = (
        _safe_series(df, "Short term borrowings", 0)
        + _safe_series(df, "Long term borrowings", 0)
    )

    with np.errstate(divide="ignore", invalid="ignore"):
        indic_rate = (fin_cost_pct * turnover) / total_debt.replace(0, np.nan) * 100

    indic_rate = indic_rate.round(2)
    df["Indicative Interest Rate (%)"] = np.where(
        indic_rate.between(7, 14), indic_rate, "DATA NA"
    )

    # ---------- Dependency % / Slab ----------
    month_col = _get_col(df, "Month")
    tofu_col = _get_col(df, "TOFU (in lacs)")
    month_ts = pd.to_datetime(df[month_col], errors="coerce")
    df["FY"] = month_ts.map(_fy_label)
    supplier_key = _get_col(df, "PAN")

    def _extrap(series: pd.Series) -> float:
        if series.count() == 0:
            return np.nan
        first_month = month_ts.loc[series.index].dt.month.min()
        months_rem = 3 - (first_month - 4) if first_month >= 4 else 3 + (4 - first_month)
        return series.sum() + series.mean() * months_rem

    tofu_fy = (
        df.groupby([supplier_key, "FY"])[tofu_col].apply(_extrap).unstack("FY")
    )

    revenue_series = _safe_series(df, "Annual Revenue", np.nan)
    if revenue_series.isna().all():
        revenue_series = df[_get_col(df, "Turnover range")].map(_parse_slab)
    revenue_sup = revenue_series.groupby(df[supplier_key]).first()

    if not tofu_fy.empty:
        latest_fy = tofu_fy.columns.sort_values()[-1]
        dep_pct = (tofu_fy[latest_fy] / revenue_sup) * 100
        dep_df = pd.DataFrame({"Dependency %": dep_pct.round(2)})
        dep_df["Dependency Slab"] = pd.cut(
            dep_df["Dependency %"],
            bins=[0, 25, 50, 75, 100, np.inf],
            labels=["<25", "25–50", "50–75", "75–100", ">100"],
            right=False,
        )
        df = df.merge(dep_df, left_on=supplier_key, right_index=True, how="left")
    else:
        df["Dependency %"] = np.nan
        df["Dependency Slab"] = np.nan

    # ---------- Industry Benchmarks ----------
    bench_key = next(
        (c for c in df.columns if c.lower() in ("industry", "nature of business")),
        None,
    )
    if bench_key is None:
        raise KeyError("Industry / Nature-of-Business column not found.")

    wc_metrics = {
        "Current Ratio": "Avg Current Ratio",
        "Receivables Days": "Avg Receivable Days",
        "Inventory Days": "Avg Inventory Days",
        "Payable Days": "Avg Payable Days",
    }
    avail_metrics: List[str] = [
        m for m in wc_metrics if m.lower() in df.columns.str.lower()
    ]

    if avail_metrics:
        for m in avail_metrics:
            df[_get_col(df, m)] = _to_num(df[_get_col(df, m)])

        bench_df = (
            df[[bench_key] + [_get_col(df, m) for m in avail_metrics]]
            .groupby(bench_key)
            .mean(numeric_only=True)
            .reset_index()
            .rename(
                columns={_get_col(df, k): v for k, v in wc_metrics.items() if k in avail_metrics}
            )
        )
    else:
        bench_df = pd.DataFrame(columns=[bench_key])

    # ---------- Deviation % ----------
    if not bench_df.empty:
        df = df.merge(bench_df, on=bench_key, how="left")
        for metric, avg_col in wc_metrics.items():
            if metric.lower() not in df.columns.str.lower():
                continue
            mcol = _get_col(df, metric)
            acol = avg_col
            df[mcol] = _to_num(df[mcol])
            df[acol] = _to_num(df[acol])
            dev_col = f"{metric} Deviation %"
            with np.errstate(divide="ignore", invalid="ignore"):
                df[dev_col] = ((df[mcol] - df[acol]).abs() / df[acol]) * 100

    return df, bench_df


# ────────────────────────── CLI entry point ──────────────────────────
def _cli() -> None:
    root = tk.Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(
        title="Select folder containing input Excel/CSV files"
    )
    if not folder_selected:
        print("❌ No folder selected. Aborting.")
        return

    files = glob.glob(os.path.join(folder_selected, "*.xlsx")) + glob.glob(
        os.path.join(folder_selected, "*.csv")
    )
    if not files:
        print("❌ No .xlsx or .csv files found in selected folder.")
        return

    for file_path in files:
        try:
            df_in = (
                pd.read_csv(file_path)
                if file_path.lower().endswith(".csv")
                else pd.read_excel(file_path)
            )
            enriched, bench = enrich_dataframe(df_in)
            out_path = os.path.splitext(file_path)[0] + "_enriched.xlsx"

            # ------- narrow Calculations sheet -------
            calc_cols_key = ["Month", "PAN", "Supplier Name"]
            derived_cols = [
                "Cash-Rich Status",
                "Indicative Interest Rate (%)",
                "Dependency %",
                "Dependency Slab",
            ] + [c for c in enriched.columns if "Deviation %" in c]

            calc_cols = [c for c in calc_cols_key + derived_cols if c in enriched.columns]
            df_calc = (
                enriched[calc_cols]
                .drop_duplicates(subset=calc_cols_key)
                .sort_values(calc_cols_key)
            )

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                wb = writer.book
                # remove the default blank sheet
                if wb.sheetnames:
                    del wb[wb.sheetnames[0]]

                # ---- Industry Benchmarks ----
                bench.to_excel(writer, sheet_name="Industry Benchmarks", index=False)

                # ---- Calculations (with colouring) ----
                ws = wb.create_sheet("Calculations")
                for row in dataframe_to_rows(df_calc, index=False, header=True):
                    ws.append(row)

                # colour-map
                fill = {
                    "good": PatternFill(start_color="C6EFCE", end_color="16C47F", fill_type="solid"),
                    "avg": PatternFill(start_color="FFEB9C", end_color="FFD65A", fill_type="solid"),
                    "bad": PatternFill(start_color="F2DCDB", end_color="F93827", fill_type="solid"),
                }

                # locate deviation columns
                for col_idx, cell in enumerate(ws[1], start=1):
                    if "Deviation %" in str(cell.value):
                        for row_idx in range(2, ws.max_row + 1):
                            c = ws.cell(row=row_idx, column=col_idx)
                            try:
                                val = float(c.value)
                            except (TypeError, ValueError):
                                continue
                            if val <= 20:
                                c.fill = fill["good"]
                            elif val <= 50:
                                c.fill = fill["avg"]
                            elif pd.isna(val) or val == "":
                                continue
                            else:
                                c.fill = fill["bad"]

            print(
                f"✅ Processed: {os.path.basename(file_path)} → {os.path.basename(out_path)}"
            )
        except Exception as e:
            print(f"❌ Failed to process {file_path}: {e}")


# ────────────────────────── CLI Execution ──────────────────────────
if __name__ == "__main__":
    _cli()
