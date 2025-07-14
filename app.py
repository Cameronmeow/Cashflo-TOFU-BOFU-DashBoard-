import streamlit as st
import pipeline.data_pull as pull
import pipeline.data_pull_new as pull_new
import pipeline.calc_all as calc
import shutil
import os
import merge_tool
import hubspot_clean as contact_dedup_tool
import company_metrics as cm
import dashboard_page as dash 
import dashboard_page_new as dash_new_v2
from io import BytesIO
import pandas as pd
import pipeline.invoice_data_pull as inv
import datetime
from company_metrics import _cli  # or expose a save helper instead
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

def excel_with_colours(enriched_df: pd.DataFrame, bench_df: pd.DataFrame) -> BytesIO:
    """
    Return an in-memory XLSX where every '… Deviation %' cell is coloured:
    ≤20 → green, 20–50 → yellow, >50 → red.
    """
    wb = Workbook()
    wb.remove(wb.active)            # kill default blank sheet

    # Benchmarks sheet
    ws_bench = wb.create_sheet("Industry Benchmarks")
    for row in dataframe_to_rows(bench_df, index=False, header=True):
        ws_bench.append(row)

    # Calculations sheet
    ws_calc = wb.create_sheet("Calculations")
    for row in dataframe_to_rows(enriched_df, index=False, header=True):
        ws_calc.append(row)

    # pastel fills
    fill_good = PatternFill("solid", fgColor="C6EFCE")   # green
    fill_avg  = PatternFill("solid", fgColor="FFEB9C")   # yellow
    fill_bad  = PatternFill("solid", fgColor="F2DCDB")   # red

    # colour every Deviation % column
    for col_idx, header_cell in enumerate(ws_calc[1], 1):
        if "Deviation %" in str(header_cell.value):
            for row_idx in range(2, ws_calc.max_row + 1):
                cell = ws_calc.cell(row=row_idx, column=col_idx)
                try:
                    val = float(cell.value)
                except (TypeError, ValueError):
                    continue
                cell.fill = (
                    fill_good if val <= 20
                    else fill_avg if val <= 50
                    else fill_bad
                )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
# import new_logic
# ─────────────────────────────────────────────
# 🌟 Page Setup
# ─────────────────────────────────────────────
st.set_page_config(page_title="CashFlo TOFU/BOFU Dashboard", layout="wide")

# ─────────────────────────────────────────────
# 🧭 Sidebar Navigation Setup
# ─────────────────────────────────────────────
st.sidebar.image("logo.webp", width=100)
st.sidebar.title("🔀 Navigation")
nav_choice = st.sidebar.radio("Go to", ["📊 TOFU BOFU Vendor Data + Cat with Revenue","📊 Vendor Category and Summary","🧹 Contact Dedup Tool","Hubspot Data CleanUp","🧩 Merge Tool", "📈 Enrichment Tool", '📁 Invoice Data Pull'])

# ─────────────────────────────────────────────
# 📊 Dashboard Logic
# ────────────────────────────────────────  ─────
if nav_choice == "📊 TOFU BOFU Vendor Data + Cat with Revenue":
    dash.render(pull_module=pull, calc_module=calc, logo_path="logo.webp")
elif nav_choice == "📊 Vendor Category and Summary":
    dash_new_v2.render(pull_module=pull_new, calc_module=calc, logo_path="logo.webp")

# … inside your nav logic:
elif nav_choice == "🧹 Contact Dedup Tool":
    st.title("🧹 Contact Dedup Tool")
    contact_dedup_tool.render_page()
# ─────────────────────────────────────────────
# 🧩 Merge Tool Logic
# ─────────────────────────────────────────────
elif nav_choice == "🧩 Merge Tool":
    st.title("🧩 Excel Merge Tool")
    merge_tool.render_page()

# ─────────────────────────────────────────────
# 📈 Enrichment Tool Logic
# ─────────────────────────────────────────────
elif nav_choice == "📈 Enrichment Tool":
    st.title("📈 Enrichment: Cash-Rich & Indicative Rate")

    uploaded_file = st.file_uploader("📁 Upload Excel or CSV file for enrichment", type=["xlsx", "csv"])
    
    if uploaded_file:
        st.success("✅ File uploaded successfully!")

        # Read file
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith("xlsx") else pd.read_csv(uploaded_file)

        # Run enrichment
        try:
            enriched_df, bench_df = cm.enrich_dataframe(df)
            st.success("✅ Enrichment completed!")

                # … after enrichment finishes
            
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                enriched_df.to_excel(w, sheet_name="Enriched Data", index=False)
                bench_df.to_excel(w, sheet_name="Industry Benchmarks", index=False)
            buf.seek(0)
            buf = excel_with_colours(enriched_df, bench_df)

            st.download_button(
                "⬇️ Download Enriched + Benchmark (coloured)",
                data=buf,
                file_name="enriched_with_benchmarks.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            
        except Exception as e:
            st.error(f"❌ Error processing file: {e}")

# elif nav_choice == "Test":
#     st.title("Test Page")
#     st.write("This is a test page for new logic.")
    
#     # Example usage of new_logic module
#     # new_logic.run(3, out_dir="Output")
#     try:
#         # result = new_?logic.run_test_logic(months_back=3, out_dir="Output")
#         excel_path = new_logic.run(months_back=3)
#         st.success(f"Test logic ran successfully: {excel_path}")
#     except Exception as e:
#         st.error(f"Error running test logic: {e}")
    
    
elif nav_choice == "📁 Invoice Data Pull":
    st.header("📁 Invoice-level Data Extract")

    # --- Date Range Picker ---
    date_range = st.date_input(
        "Select Date Range",
        value=(
            datetime.date.today().replace(day=1),
            datetime.date.today()
        ),
        format="YYYY-MM-DD",
        help="Choose the start and end date for the extract"
    )

    if not isinstance(date_range, tuple) or len(date_range) != 2:
        st.error("Please select a valid start and end date.")
    else:
        from_date, to_date = date_range

        # --- Granularity ---
        granularity_sel = st.radio(
            "Granularity (how to bucket rows)", 
            ["daily", "weekly"], 
            horizontal=True
        )

        # --- Date Type Filter ---
        date_type_sel = st.selectbox(
            "Date column used for filtering / bucketing",
            options=["i.createdAt", "epri.activatedOn", "epri.toBeClearedOnUtc"],
            format_func=lambda x: {
                "i.createdAt"        : "Invoice Created At (i.createdAt)",
                "epri.activatedOn"   : "EPR Activated On (epri.activatedOn)",
                "epri.toBeClearedOnUtc":
                    "EPR To-Be-Cleared On (epri.toBeClearedOnUtc)"
            }[x]
        )

        st.caption(f"➡️ Pulling data between **{from_date}** and **{to_date}**, bucketed **{granularity_sel}**, using **{date_type_sel}**")
        
        # --- Button ---
        if st.button("🚀 Run invoice-level pull"):
            with st.spinner("Running query – this may take a minute…"):
                df_inv = inv.run_invoice_pull(
                    from_date   = from_date,
                    to_date     = to_date,
                    granularity = granularity_sel,
                    date_type   = date_type_sel
                )
                # for col in df_inv.select_dtypes(include=["datetimetz"]).columns:
                #     df_inv[col] = df_inv[col].dt.tz_convert(None)

            st.success(f"✅ Pulled {len(df_inv):,} rows")
            st.dataframe(df_inv.head())
            
            # --- Download Excel ---
            out_name = f"invoice_metrics_{from_date}_{to_date}_{granularity_sel}.xlsx"
            with open(f"Output/{out_name}", "rb") as f:
                st.download_button(
                    "⬇️ Download Excel",
                    data=f,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )